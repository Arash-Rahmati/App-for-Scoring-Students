#for elementary worksheets
import pandas as pd                     #to create a neat DataFrame
import matplotlib.pyplot as pyplot      #to create charts when needed
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import arabic_reshaper
from bidi.algorithm import get_display
import numpy as np


#read the list of names with theirs rows next to them from excel
#for my dataframe
dfnames=[]
dftotal=[]
dflistening=[]
dfreading=[]
dfspeaking=[]
dfwriting=[]
dfworkbook=[]
dfindex=[]
#for my dataframe

wb = load_workbook('Class.xlsx')

stringtemp= input('Which of your classes do your want? 1, 2, 3, 4, 5, 6\n')   #the first class of the week is number 1
ws = wb[str(stringtemp)]
addresss=r'C:\Users\arash\Desktop\python\Class'+stringtemp   #data of the session will be saved here later

j=2
while ws['AY'+str(j)].value!=None:

        examplestring=str(ws['AY'+str(j)].value)
        # for dataframe
        dfnames.append(examplestring)
        dflistening.append(0)
        dfreading.append(0)
        dfspeaking.append(0)
        dfwriting.append(0)
        dfworkbook.append(0)
        dftotal.append(0)
        dfindex.append(j)
        j+=1
        # for dataframe



#create the datafram
mydict={"Names":dfnames,"Total": dftotal, "Listening":dflistening,"Speaking":dfspeaking,"Reading":dfreading,"Writing":dfwriting,"Workbook":dfworkbook}
mydataframe=pd.DataFrame(mydict)
mydataframe.index=pd.Series(dfindex)
Number_of_Students = mydataframe.shape[0]

#make the required changes
while 1<2:
   ch=input("Press 1: for Listening     Press 5: for Print\n"+"Press 2: for Speaking      Press 6: for Ordered Print   Print  9: For Graph\n"+"Press 3: for Reading       Press 7: for Save and Exit   Press 10: For Note\n"+"Press 4: for Writing       Press 8: for Workbook        Press 11: Describtive Score\n")


   if int(ch)==1:
       temp = list(map(int, input("Enter Data: ").split())) #this reads all in one line separate by space
       for k in range(len(temp)):
           if 1<temp[k]<22:
            #for dataframe
            mydataframe.loc[temp[k],'Listening']+=1
            mydataframe.loc[temp[k],'Total']+=1
            #for dataframe
            mytemp='W'+str(temp[k])
            ws[mytemp].value+=1
           else: 
               print("\n !!! A number was OUt of RanGe. Check it out!!! \n")


   elif int(ch)==2:
       temp = list(map(int, input("Enter Data: ").split())) #this reads all in one line separate by space
       for k in range(len(temp)):
           if 1<temp[k]<22:
            #for dataframe
            mydataframe.loc[temp[k],'Speaking']+=1
            mydataframe.loc[temp[k],'Total']+=1
            #for dataframe
            mytemp='AB'+str(temp[k])
            ws[mytemp].value+=1
           else: 
               print("\n !!! A number was OUt of RanGe. Check it out!!! \n")

   elif int(ch)==3:
       temp = list(map(int, input("Enter Data: ").split())) #this reads all in one line separate by space
       for k in range(len(temp)):
           if 1<temp[k]<22:
            #for dataframe
            mydataframe.loc[temp[k],'Reading']+=1
            mydataframe.loc[temp[k],'Total']+=1
            #for dataframe
            mytemp = 'AG' + str(temp[k])
            ws[mytemp].value += 1
           else: 
               print("\n !!! A number was OUt of RanGe. Check it out!!! \n")

   elif int(ch)==4:
       temp = list(map(int, input("Enter Data: ").split())) #this reads all in one line separate by space
       for k in range(len(temp)):
           if 1<temp[k]<22:
            #for dataframe
            mydataframe.loc[temp[k],'Writing']+=1
            mydataframe.loc[temp[k],'Total']+=1
            #for dataframe
            mytemp='AL'+str(temp[k])
            ws[mytemp].value+=1
           else: 
               print("\n !!! A number was OUt of RanGe. Check it out!!! \n")

   elif int(ch)==5:
       #printlist(student_list)
       print(mydataframe)

   elif int(ch)==6:

        print(mydataframe.sort_values(['Total'], ascending=False))

   elif int(ch)==7:
        flag=0
        check=0
        try:
            wb.save('Class.xlsx')
        except:
            print('Try closing the excel file and save again.')
            flag=-1
        
        if (flag!=-1):
          
            title = input('class and session')
            session = title[len(title)-2]+title[len(title)-1]
            title2=title+'.pdf'
            title2= os.path.join(addresss,title2)

            r=2
            while ws['B'+str(r)].value!=None:
                mydataframe.loc[r,"Names"]=ws['B'+str(r)].value
                r+=1
            mydataframe.sort_values(['Total'], ascending=False, inplace=True)

            try:
                with pd.ExcelWriter(os.path.join(addresss,'ExcelData.xlsx'),mode='a') as writer:    
                    try:
                        mydataframe.to_excel(writer, sheet_name=session)
                    except:
                        print("this name already exists.!!!")
                        check=-1    
            except:
                print('ExcelData file is open retry.')
                check=-1


            index=[]
            for i in range(Number_of_Students): index.append(i+1)
            names =[]
            scores=[]
            for i in range(Number_of_Students):  
                farsiname=ws['B'+str(i+2)].value
                farsiname=arabic_reshaper.reshape(farsiname)
                farsiname=get_display(farsiname)
                names.append(farsiname)
            for i in range(Number_of_Students): scores.append(mydataframe.loc[i+2,'Total'])
            pyplot.barh(index,scores,tick_label=names)

            
            pyplot.subplots_adjust(0.3,bottom=0.05, right=None, top=0.92, wspace=None, hspace=None)
            pyplot.title("Students' Activity in Session "+session)
            if(check==0): pyplot.savefig(title2)

            if(check==0): break

   elif int(ch)==8:
       temp = list(map(int, input("Enter Data: ").split())) #this reads all in one line separate by space
       for k in range(len(temp)):
           if 1<temp[k]<22:
            #for dataframe
            mydataframe.loc[temp[k],'Workbook']+=1
            mydataframe.loc[temp[k],'Total']+=1
            #for dataframe
            mytemp='AQ'+str(temp[k])
            ws[mytemp].value+=1
           else: 
               print("\n !!! A number was OUt of RanGe. Check it out!!! \n")




   elif int(ch)==9:
        index=[]
        for i in range(Number_of_Students): index.append(i+1)
        names =[]
        scores=[]
        for i in range(Number_of_Students): 
            farsiname=ws['B'+str(i+2)].value
            farsiname=arabic_reshaper.reshape(farsiname)
            farsiname=get_display(farsiname)
            names.append(farsiname)
        for i in range(Number_of_Students): scores.append(mydataframe.loc[i+2,'Total'])
        pyplot.barh(index,scores,tick_label=names)
        pyplot.xlabel=('learners')
        pyplot.ylabel=("today's score")
        pyplot.subplots_adjust(0.3,bottom=0.05, right=None, top=0.92, wspace=None, hspace=None)
        pyplot.show()

   elif int(ch)==10:
        
        ind=int(input('Which student:  //numerical Data Required or the App would crash '))
        if(2<=ind<=22):                
            ind=str(ind)
            currentstring=ws['BB'+ind].value
            sess=input('What Session is it: ')
            note=input('What is your note for '+ws['AY'+ind].value +' : ')    
            newstring=currentstring+sess+':'+note+' # \n'
            ws['BB'+ind].value=newstring
            print('Note was added thanks.') 

        else: 
            print('invalid input')      

   elif int(ch)==11:
        added=0
        descriptive=['BP','BQ','BR','BS','BT','BU','BV']
        ind=int(input('Which student Descriptive Score:  //numerical Data Required or the App would crash '))
        if(2<=ind<=22):                
            note=input('What is your Descriptive Score for '+ws['AY'+str(ind)].value +' : ') 
            for i in range(7):
                if(ws[descriptive[i]+str(ind)].value==None): 
                    ws[descriptive[i]+str(ind)].value=int(note)
                    added=1
                    break
            if(added==0):
                arr=np.array([])
                for i in range(7):
                        arr=np.append(arr,ws[descriptive[i]+str(ind)].value)
                minn=np.min(arr)
                for i in range(7):
                        if(ws[descriptive[i]+str(ind)].value==minn): 
                            ws[descriptive[i]+str(ind)].value=int(note)
                            break
            print('Descriptive Score was added/updated thanks.') 



        else: 
            print('invalid input') 




print('Have a Good Day ;) You can leave now...')

#wb.save('Basic.xlsx') use it as an option to end and save